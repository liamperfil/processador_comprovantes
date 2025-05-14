import os
import re
import shutil
import platform
from datetime import datetime
from PyPDF2 import PdfReader
from openpyxl import load_workbook

arquivo_xlsx = "pagamentos.xlsx"
planilhas_clientes = ["amoedo", "beatriz", "cavalcante", "dsouza", "julia"]

hoje_str = datetime.now().strftime("%Y-%m-%d")
log_dir = "logs"
os.makedirs(log_dir, exist_ok=True)
log_path = os.path.join(log_dir, f"log_{hoje_str}.txt")

def registrar_log(mensagem):
    print(mensagem)
    with open(log_path, "a", encoding="utf-8") as log_file:
        log_file.write(mensagem + "\n")

def normalizar_codigo_barras(texto):
    padroes = [
        r"\b(\d{5}\s\d{5}\s\d{5}\s\d{6}\s\d{5}\s\d{6}\s\d\s\d{14})\b",
        r"\b(\d{47,48})\b",
        r"(?:\d{11}-\d{1}\s+){3}\d{11}-\d{1}"
    ]
    texto_limpo = texto.replace("\n", " ")
    for padrao in padroes:
        match = re.search(padrao, texto_limpo)
        if match:
            codigo = re.sub(r"[\s\-]", "", match.group(0))
            return codigo[:47]
    return None

def extrair_data_pagamento(texto, cliente):
    if cliente == "julia":
        match = re.search(r"(\d{2}/\d{2}/\d{4})\s*data de d[√™e]bito", texto, re.IGNORECASE)
        if match:
            return match.group(1)
    else:
        match = re.search(r"(?:data(?: do pagamento| de d[√™e]bito| agendamento| de agendamento| agendamento)[\s:]*)(\d{2}/\d{2}/\d{4})", texto, re.IGNORECASE)
        if match:
            return match.group(1)
    return None

def extrair_valor_cobrado(texto, cliente):
    if cliente == "julia":
        padroes_julia = [
            r"R?\$?\s*(\d{1,3}(?:\.\d{3})*,\d{2})\s*valor total",
            r"R?\$?\s*(\d{1,3}(?:\.\d{3})*,\d{2})\s*valor do pagamento",
            r"R?\$?\s*(\d{1,3}(?:\.\d{3})*,\d{2})\s*valor"
        ]
        for padrao in padroes_julia:
            match = re.search(padrao, texto, re.IGNORECASE)
            if match:
                return match.group(1).replace(".", "").replace(",", ".")
    else:
        padroes_outros = [
            r"(?:valor total|valor cobrado)[\s:]*R?\$?\s*(\d{1,3}(?:\.\d{3})*,\d{2})",
            r"R?\$?\s*(\d{1,3}(?:\.\d{3})*,\d{2})"
        ]
        for padrao in padroes_outros:
            match = re.search(padrao, texto, re.IGNORECASE)
            if match:
                return match.group(1).replace(".", "").replace(",", ".")
    return None

def identificar_cliente(texto):
    texto = texto.lower()
    match = re.search(r"(cliente|empresa)\s*:\s*(.+)", texto, re.IGNORECASE)
    if match:
        linha_cliente = match.group(2).strip().lower()
        if "amoedo" in linha_cliente:
            return "amoedo"
        elif "beatriz" in linha_cliente:
            return "beatriz"
        elif "cavalcante" in linha_cliente:
            return "cavalcante"
        elif "dsouza" in linha_cliente or "souza" in linha_cliente:
            return "dsouza"
        elif "transportes" in linha_cliente or "julia" in linha_cliente:
            return "julia"
    if "amoedo" in texto:
        return "amoedo"
    elif "beatriz" in texto:
        return "beatriz"
    elif "cavalcante" in texto:
        return "cavalcante"
    elif "dsouza" in texto or "souza" in texto:
        return "dsouza"
    elif "transportes" in texto or "julia" in texto:
        return "julia"
    return None

def criar_backup_planilha():
    now = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = os.path.join(log_dir, f"pagamentos_backup_{now}.xlsx")
    shutil.copy2(arquivo_xlsx, backup_path)
    registrar_log(f"[INFO] Backup criado: {backup_path}")

def processar_pdfs():
    arquivos_pdf = [arq for arq in os.listdir() if arq.lower().endswith(".pdf")]
    if not arquivos_pdf:
        registrar_log("[INFO] Nenhum arquivo PDF encontrado no diret√≥rio.")
        return

    criar_backup_planilha()
    wb = load_workbook(arquivo_xlsx)

    for arquivo in arquivos_pdf:
        try:
            reader = PdfReader(arquivo)
            texto = "\n".join(page.extract_text() or '' for page in reader.pages)

            cliente = identificar_cliente(texto)
            if not cliente:
                registrar_log(f"[WARNING] {arquivo} - Cliente n√£o identificado.")
                continue

            if cliente not in wb.sheetnames:
                registrar_log(f"[WARNING] {arquivo} - Planilha '{cliente}' n√£o encontrada.")
                continue

            plan = wb[cliente]
            codigo_barras = normalizar_codigo_barras(texto)
            if not codigo_barras:
                registrar_log(f"[WARNING] {arquivo} - C√≥digo de barras n√£o encontrado no PDF.")
            data_pagamento = extrair_data_pagamento(texto, cliente)
            valor_cobrado = extrair_valor_cobrado(texto, cliente)

            cabecalho = {str(cell.value).strip().lower(): idx for idx, cell in enumerate(plan[1]) if cell.value}
            campos = ["id", "pagamento", "vencimento", "codigo de barras", "status", "origem"]
            if not all(c in cabecalho for c in campos):
                registrar_log(f"[ALERT] {arquivo} - Cabe√ßalho incompleto na planilha '{cliente}'.")
                continue

            linha_encontrada = None
            if codigo_barras:
                linhas_possiveis = []
                for row in plan.iter_rows(min_row=2):
                    celula_cod = row[cabecalho["codigo de barras"]]
                    if celula_cod.value and codigo_barras == str(celula_cod.value).strip():
                        if not row[cabecalho["pagamento"]].value:
                            linhas_possiveis.append(row)

                if linhas_possiveis:
                    linha_encontrada = linhas_possiveis[0]
                elif any(str(row[cabecalho["codigo de barras"]].value).strip() == codigo_barras for row in plan.iter_rows(min_row=2)):
                    registrar_log(f"[WARNING] {arquivo} - C√≥digo de barras encontrado, mas todas as linhas j√° foram processadas.")

            if not linha_encontrada:
                os.system("cls" if platform.system() == "Windows" else "clear")

                print(f"---------------- C√≥digo de barras n√£o encontrado na planilha {cliente} ----------------")
                print(f"\nInforma√ß√µes extra√≠das do PDF:")
                print(f"Arquivo: {arquivo}.")
                texto_limpo = re.sub(r"\s+", " ", texto).strip()
                resumo = texto_limpo[:800] + "..." if len(texto_limpo) > 800 else texto_limpo
                print(f"Cliente: {cliente}")
                print(f"Data de pagamento: {data_pagamento or 'N/A'}")
                print(f"Valor cobrado: {valor_cobrado or 'N/A'}")
                print(f"C√≥digo de barras: {codigo_barras or 'N/A'}")
                print(f"\nResumo do conte√∫do:\n{resumo}")
                try:
                    id_manual = int(input("\nDigite o ID correspondente ou pressione Enter para ignorar: ").strip())
                except ValueError:
                    if not input("\nID inv√°lido. Pressione Enter para ignorar ou digite qualquer coisa para tentar novamente: "):
                        continue  # Ignora se Enter for pressionado
                    else:
                        continue  # Volta ao in√≠cio do loop para tentar novamente

                for row in plan.iter_rows(min_row=2):
                    plan_id_value = row[cabecalho["id"]].value
                    if isinstance(plan_id_value, (int, float)):
                        plan_id = int(plan_id_value)  # Converter para int se for int ou float
                    else:
                        plan_id = int(str(plan_id_value).strip())  # Converter para int ap√≥s limpar a string
                    if plan_id == id_manual:
                        if not row[cabecalho["pagamento"]].value:
                            linha_encontrada = row
                        break

            if linha_encontrada:
                if not data_pagamento:
                    vencimento_val = linha_encontrada[cabecalho["vencimento"]].value
                    data_pagamento = vencimento_val if vencimento_val else None

                if data_pagamento:
                    if isinstance(data_pagamento, str):
                        data_pagamento = datetime.strptime(data_pagamento, "%d/%m/%Y")
                    linha_encontrada[cabecalho["pagamento"]].value = data_pagamento
                    linha_encontrada[cabecalho["status"]].value = "Pago"
                    linha_encontrada[cabecalho["origem"]].value = "Bradesco" if cliente == "julia" else "Banco do Brasil"

                    if valor_cobrado:
                        try:
                            valor_float = float(valor_cobrado)
                            if "valor cobrado" in cabecalho:
                                linha_encontrada[cabecalho["valor cobrado"]].value = valor_float
                            elif "valor do documento" in cabecalho:
                                linha_encontrada[cabecalho["valor do documento"]].value = valor_float
                        except ValueError:
                            registrar_log(f"[WARNING] {arquivo} - Valor cobrado inv√°lido: {valor_cobrado}")

                    id_valor_raw = linha_encontrada[cabecalho["id"]].value
                    id_valor = str(int(float(id_valor_raw))) if isinstance(id_valor_raw, (float, int)) else str(id_valor_raw).lstrip("0")
                    data_para_nome = data_pagamento.strftime("%Y%m%d") if data_pagamento else ""  # Formatando a data
                    novo_nome = f"{cliente}_{data_para_nome}_{id_valor}.pdf" if data_para_nome else f"{cliente}_{id_valor}.pdf" # Criando o novo nome
                    
                    ano_mes = data_pagamento.strftime("%Y-%m") if data_pagamento else datetime.now().strftime("%Y-%m") # Usar a data atual se data_pagamento for None
                    destino = os.path.join("comprovantes", ano_mes)
                    os.makedirs(destino, exist_ok=True)
                    destino = os.path.join(destino, novo_nome) # Caminho completo com o nome do arquivo
                    
                    shutil.move(arquivo, destino)
                    registrar_log(f"[INFO] {arquivo} ‚Üí {novo_nome} - Processado com sucesso.")
                else:
                    registrar_log(f"[WARNING] {arquivo} - Data de pagamento e vencimento n√£o encontradas.")
            elif not data_pagamento:
                registrar_log(f"[WARNING] {arquivo} - Data de pagamento n√£o encontrada.")
            else:
                registrar_log(f"[WARNING] {arquivo} - Nenhuma linha correspondente encontrada com ID informado.")

        except Exception as e:
            registrar_log(f"[ALERT] {arquivo} - Erro durante o processamento: {e}")

    wb.save(arquivo_xlsx)
    registrar_log("[INFO] üìÑ Planilha salva com sucesso.")

if __name__ == "__main__":
    processar_pdfs()
